#!/usr/bin/env python3
"""
AVL-DRIVE Heatmap Tool — Python Edition (Standalone)
=====================================================
A 1:1 standalone Python replica of the Excel VBA-based AVL-DRIVE Heatmap Tool
(version 5.1). Completely independent — does NOT require the original .xlsm file.

Users upload their own data files:
  - Data Transfer Sheet — CSV or Excel with AVL-DRIVE scores per vehicle
  - Sheet1              — Excel (.xlsx) with benchmark data and coloured P1 dots

All reference data (Mapping Sheet, AVL-Odriv Mapping, HeatMap Template) is
embedded directly in this file.

Provides a Streamlit GUI with the same six buttons as the Excel ribbon:
  1. HeatMap          — Refresh heatmap from Data Transfer Sheet
  2. Reset            — Restore HeatMap Sheet from HeatMap Template
  3. Evaluation       — Evaluate AVL statuses with car selection
  4. Suboperation     — Write coloured dots to HeatMap Sheet
  5. Op Mode Status   — Aggregate group statuses
  6. Export           — Export visible data as XLSX

Usage:
    streamlit run avl_heatmap_tool.py
"""

from __future__ import annotations

import copy
import io
import os
import sys
import textwrap
from collections import OrderedDict
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import streamlit as st

# ═══════════════════════════════════════════════════════════════════════════════
#  EMBEDDED REFERENCE DATA
# ═══════════════════════════════════════════════════════════════════════════════

# Mapping Sheet — 58 entries (op code -> canonical operation name)
MAPPING_SHEET_DATA: List[Tuple[int, str]] = [
    (10000000, "AVL-DRIVE Rating"),
    (10100000, "Drive away"),
    (10101300, "Creep"),
    (10101100, "Standing start"),
    (10102400, "Rolling start"),
    (10120000, "Acceleration"),
    (10120100, "Full load"),
    (10120200, "Constant load"),
    (10120300, "Load increase"),
    (10120900, "Load decrease"),
    (10030000, "Tip in"),
    (10030100, "At deceleration"),
    (10030200, "At constant speed / acceleration"),
    (10040000, "Tip out"),
    (10040300, "At constant speed / acceleration"),
    (10040400, "At deceleration"),
    (10070100, "Transition to constant speed"),
    (10070000, "Deceleration"),
    (10070500, "Without brake"),
    (10071000, "Constant brake"),
    (10090000, "Gear shift"),
    (10092300, "Power-on upshift"),
    (10092500, "Tip out upshift"),
    (10098200, "Tip in upshift"),
    (10098400, "Load reversal upshift"),
    (10092100, "Coast / brake-on upshift"),
    (10093200, "Power-on downshift"),
    (10098100, "Tip out downshift"),
    (10093100, "Kick down / tip in downshift"),
    (10098300, "Load reversal downshift"),
    (10093400, "Coast / brake-on downshift"),
    (10097800, "Maneuvering"),
    (10097900, "Selector lever change"),
    (10080000, "Constant speed"),
    (10080200, "Without load"),
    (10080100, "Constant load"),
    (10010000, "Idle"),
    (10011000, "Vehicle stationary"),
    (10010200, "Air conditioning on / off"),
    (10010700, "Transition to idle"),
    (10015200, "Rev-up"),
    (10020000, "Engine start"),
    (10020100, "Manual start"),
    (10020200, "Auto start vehicle stationary"),
    (10020300, "Auto start vehicle moving"),
    (10140000, "Engine shut off"),
    (10140600, "Manual stop"),
    (10140700, "Auto stop"),
    (10460000, "TCC control"),
    (10467300, "Converter controlled slip"),
    (10467200, "Converter lock up"),
    (10467500, "Converter release"),
    (10430000, "Cylinder deactivation"),
    (10431300, "Cylinder deactivation"),
    (10431400, "Cylinder reactivation"),
    (10450000, "Vehicle stationary"),
    (10451400, "Vehicle stop"),
    (10451500, "Vehicle at standstill"),
]

# AVL-Odriv Mapping — maps AVL op codes to Sheet1 sub-operation names
AVL_ODRIV_MAPPING_DATA: List[Tuple[int, str]] = [
    (10000000, "AVL-DRIVE Rating"),
    (10100000, "Drive away"),
    (10101300, "Creep"),
    (10101300, "Drive Away Creep Eng On"),
    (10101300, "Drive Away Creep Eng On - Cold"),
    (10101300, "Drive Away Creep Eng Off"),
    (10101100, "Standing start"),
    (10101100, "DASS Eng On"),
    (10101100, "DASS Eng On - Cold"),
    (10101100, "DASS Eng Off quick"),
    (10101100, "DASS Eng Off slow"),
    (10101100, "DASS - Eng Off - COM"),
    (10101100, "DASS - Extended Eng Off"),
    (10101100, "DASS"),
    (10102400, "Rolling start"),
    (10102400, "DA Rolling Start"),
    (10120000, "Acceleration"),
    (10120100, "Full load"),
    (10120200, "Constant load"),
    (10120200, "Accel Cst Load"),
    (10120200, "Accel Cst Load - Cold"),
    (10120300, "Load increase"),
    (10120300, "Accel Load Increase"),
    (10120900, "Load decrease"),
    (10120900, "Accel Load Decrease"),
    (10030000, "Tip in"),
    (10030100, "At deceleration"),
    (10030100, "Tip in at deceleration"),
    (10030200, "At constant speed / acceleration"),
    (10030200, "Tip in at constant speed"),
    (10040000, "Tip out"),
    (10040300, "Tip Out At Constant Speed"),
    (10040300, "At constant speed / acceleration"),
    (10040400, "Tip Out After Acceleration"),
    (10040400, "At deceleration"),
    (10070000, "Deceleration"),
    (10070100, "Decel Trans to Cst Spd"),
    (10070100, "Decel - Trans to Cst Spd - Cold"),
    (10070500, "Without brake"),
    (10070500, "Decel Without Brake"),
    (10070500, "Decel Without Brake - Cold"),
    (10071000, "Constant brake"),
    (10071000, "Decel Cst Brake"),
    (10071000, "Decel Cst Brake - Cold"),
    (10090000, "Gear shift"),
    (10092300, "Power-on upshift"),
    (10092300, "Power-on upshift"),
    (10092300, "Power-on upshift Cold"),
    (10092500, "Tip out upshift"),
    (10098200, "Tip in upshift"),
    (10098400, "Load reversal upshift"),
    (10092100, "Coast / brake-on upshift"),
    (10093200, "Power-on downshift"),
    (10098100, "Tip out downshift"),
    (10093100, "Kick down / tip in downshift"),
    (10093100, "(PT) KD - tip in downshift"),
    (10093100, "(TO) KD - tip in downshift"),
    (10098300, "Load reversal downshift"),
    (10093400, "Coast / brake-on downshift"),
    (10093100, "Coast-brake-on downshift Cold"),
    (10097800, "Maneuvering"),
    (10097800, "Maneuvering - Cold"),
    (10097800, "Maneuvering with throttle"),
    (10097900, "Selector lever change"),
    (10097900, "Lever change"),
    (10080000, "Constant speed"),
    (10080200, "Without load"),
    (10080200, "Cst Speed Without Load"),
    (10080200, "Cst Speed Without Load - Cold"),
    (10080100, "Constant load"),
    (10080100, "Cst Speed Cst Load"),
    (10080100, "Cst Speed Cst Load - Cold"),
    (10010000, "Idle"),
    (10011000, "Vehicle stationary"),
    (10010200, "Air conditioning on / off"),
    (10010200, "Idle Air Cond On-Off"),
    (10010700, "Transition to idle"),
    (10015200, "Rev-up"),
    (10020000, "Engine start"),
    (10020100, "Manual start"),
    (10020200, "Auto start vehicle stationary"),
    (10020300, "Auto start vehicle moving"),
    (10140000, "Engine shut off"),
    (10140600, "Manual stop"),
    (10140700, "Auto stop"),
    (10460000, "TCC control"),
    (10467300, "Converter controlled slip"),
    (10467200, "Converter lock up"),
    (10467500, "Converter release"),
    (10430000, "Cylinder deactivation"),
    (10431300, "Cylinder deactivation"),
    (10431400, "Cylinder reactivation"),
    (10450000, "Vehicle stationary"),
    (10450000, "Idle Vehicle Stationary"),
    (10450000, "Idle Vehicle Stationary - Cold"),
    (10451400, "Vehicle stop"),
    (10451400, "Vehicle Stop"),
    (10451400, "Vehicle Stop - Cold"),
    (10451500, "Vehicle at standstill"),
]

# HeatMap Template row definitions: (row_num, op_code, name, is_bold_group_header)
HEATMAP_TEMPLATE_ROWS: List[Tuple[int, int, str, bool]] = [
    (4, 10000000, "AVL-DRIVE Rating", False),
    (5, 10100000, "Drive away", True),
    (6, 10101300, "Creep", False),
    (7, 10101100, "Standing start", False),
    (8, 10102400, "Rolling start", False),
    (9, 10120000, "Acceleration", True),
    (10, 10120100, "Full load", False),
    (11, 10120200, "Constant load", False),
    (12, 10120300, "Load increase", False),
    (13, 10120900, "Load decrease", False),
    (14, 10030000, "Tip in", True),
    (15, 10030100, "At deceleration", False),
    (16, 10030200, "At constant speed / acceleration", False),
    (17, 10040000, "Tip out", True),
    (18, 10040300, "At constant speed / acceleration", False),
    (19, 10040400, "At deceleration", False),
    (20, 10070000, "Deceleration", True),
    (21, 10070500, "Without brake", False),
    (22, 10070100, "Transition to constant speed", False),
    (23, 10071000, "Constant Brake", False),
    (24, 10090000, "Gear shift", True),
    (25, 10092300, "Power-on upshift", False),
    (26, 10092500, "Tip out upshift", False),
    (27, 10098200, "Tip in upshift", False),
    (28, 10098400, "Load reversal upshift", False),
    (29, 10092100, "Coast / brake-on upshift", False),
    (30, 10093200, "Power-on downshift", False),
    (31, 10098100, "Tip out downshift", False),
    (32, 10093100, "Kick down / tip in downshift", False),
    (33, 10098300, "Load reversal downshift", False),
    (34, 10093400, "Coast / brake-on downshift", False),
    (35, 10097800, "Maneuvering", False),
    (36, 10097900, "Selector lever change", False),
    (37, 10080000, "Constant speed", True),
    (38, 10080200, "Without load", False),
    (39, 10080100, "Constant load", False),
    (40, 10010000, "Idle", True),
    (41, 10011000, "Vehicle stationary", False),
    (42, 10010200, "Air conditioning on / off", False),
    (43, 10010700, "Transition to idle", False),
    (44, 10015200, "Rev-up", False),
    (45, 10020000, "Engine start", True),
    (46, 10020100, "Manual start", False),
    (47, 10020200, "Auto start vehicle stationary", False),
    (48, 10020300, "Auto start vehicle moving", False),
    (49, 10140000, "Engine shut off", True),
    (50, 10140600, "Manual stop", False),
    (51, 10140700, "Auto stop", False),
    (52, 10460000, "TCC control", True),
    (53, 10467300, "Converter controlled slip", False),
    (54, 10467200, "Converter lock up", False),
    (55, 10467500, "Converter release", False),
    (56, 10430000, "Cylinder deactivation", True),
    (57, 10431300, "Cylinder deactivation", False),
    (58, 10431400, "Cylinder reactivation", False),
    (59, 10450000, "Vehicle stationary", True),
    (60, 10451400, "Vehicle stop", False),
    (61, 10451500, "Vehicle at standstill", False),
]

# ─── Sheet name constants ────────────────────────────────────────────────────

SHEET_T = "HeatMap Sheet"
SHEET_S = "Data Transfer Sheet"
TEMPLATE_SHEET = "HeatMap Template"
SHEET1 = "Sheet1"
EVAL_RESULTS = "Evaluation Results"

ANCHOR_TEXT = "Operation Modes"
TARGET_VEHICLE_HEADER = "Target Vehicle"
TESTED_VEHICLE_HEADER = "Tested Vehicle"

HIDE_IDS_COLA = True
DELETE_EMPTY = False

# Vehicle columns in HeatMap Sheet (1-indexed)
VEHICLE_COLS = [4, 6, 8, 10, 12, 14, 16]     # D F H J L N P
SEPARATOR_COLS = [5, 7, 9, 11, 13, 15, 17]   # E G I K M O Q

# Excel indexed colour palette (standard)
_INDEXED_COLORS = {
    0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00", 4: "0000FF",
    5: "FFFF00", 6: "FF00FF", 7: "00FFFF", 8: "000000", 9: "FFFFFF",
    10: "FF0000", 11: "00FF00", 12: "0000FF", 13: "FFFF00", 14: "FF00FF",
    15: "00FFFF", 16: "800000", 17: "008000", 18: "000080", 19: "808000",
    20: "800080", 21: "008080", 22: "C0C0C0", 23: "808080",
}

# Fill for status cells
FILL_GREEN = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SUMMARY = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")
FONT_BOLD_WHITE = Font(color="FFFFFF", bold=True)

BULLET = "\u25CF"  # ●

CAR_DATA_START_COL = 8  # Column H in Sheet1


# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITY HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _cell_val(ws, row: int, col: int) -> Any:
    """Return cell value (None-safe)."""
    return ws.cell(row, col).value


def _to_float(v: Any) -> float:
    """Convert to float; return 0.0 for non-numeric."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _is_numeric(v: Any) -> bool:
    """Return True if *v* can be interpreted as a number."""
    if v is None or v == "":
        return False
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def _trim(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _resolve_font_color_hex(cell) -> str:
    """Return 6-char uppercase hex of the font colour, or 'FFFFFF' (white)."""
    fc = cell.font.color
    if fc is None:
        return "FFFFFF"
    if fc.type == "rgb" and fc.rgb:
        h = str(fc.rgb)
        if len(h) == 8:
            h = h[2:]  # strip alpha
        return h.upper()
    if fc.type == "indexed" and fc.indexed is not None:
        idx = fc.indexed
        if idx in _INDEXED_COLORS:
            return _INDEXED_COLORS[idx].upper()
    if fc.type == "theme":
        return "FFFFFF"
    return "FFFFFF"


def _resolve_fill_color_hex(cell) -> str:
    """Return 6-char uppercase hex of the cell fill colour, or '000000'."""
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return "000000"
    fg_color = fill.fgColor
    if fg_color.type == "rgb" and fg_color.rgb:
        h = str(fg_color.rgb)
        if len(h) == 8:
            h = h[2:]
        return h.upper()
    if fg_color.type == "indexed" and fg_color.indexed is not None:
        idx = fg_color.indexed
        if idx in _INDEXED_COLORS:
            return _INDEXED_COLORS[idx].upper()
    return "000000"


def _is_near(r: int, g: int, b: int, rt: int, gt: int, bt: int, tol: int = 45) -> bool:
    return abs(r - rt) <= tol and abs(g - gt) <= tol and abs(b - bt) <= tol


def _hex_to_rgb(h: str) -> Tuple[int, int, int]:
    h = h.lstrip("#")
    if len(h) == 8:
        h = h[2:]
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def _find_anchor(ws, text: str) -> Optional[Tuple[int, int]]:
    """Find the row, col of the anchor text (case-insensitive, whole-cell)."""
    tl = text.strip().lower()
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = _cell_val(ws, row, col)
            if v is not None and _trim(v).lower() == tl:
                return (row, col)
    return None


def _has_value(v: Any) -> bool:
    """Mirror VBA HasValue: True if v is a positive number or non-empty non-zero string."""
    if v is None:
        return False
    try:
        if isinstance(v, (int, float)):
            return float(v) > 0
        fv = float(v)
        return fv > 0
    except (ValueError, TypeError):
        s = str(v).strip()
        return s != "" and s != "0"


def _last_data_row(ws, col: int) -> int:
    """Find the last non-empty row in a column."""
    for r in range(ws.max_row, 0, -1):
        if _cell_val(ws, r, col) is not None:
            return r
    return 1


# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD INTERNAL WORKBOOK FROM UPLOADED DATA
# ═══════════════════════════════════════════════════════════════════════════════

def build_workbook_from_uploads(
    dt_wb: openpyxl.Workbook,
    s1_wb: openpyxl.Workbook,
) -> openpyxl.Workbook:
    """Create a complete internal workbook from two user-uploaded files.

    Parameters
    ----------
    dt_wb : user-uploaded Data Transfer Sheet workbook (first sheet used)
    s1_wb : user-uploaded Sheet1 workbook (first sheet used, preserves colours)

    Returns
    -------
    openpyxl.Workbook with all required sheets populated.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. Copy Sheet1 from upload (preserves P1 colours)
    _copy_sheet(s1_wb.active, wb, SHEET1)

    # 2. Copy Data Transfer Sheet from upload
    _copy_sheet(dt_wb.active, wb, SHEET_S)

    # 3. Build HeatMap Template (from embedded data)
    _build_heatmap_template(wb)

    # 4. Build HeatMap Sheet (copy of template)
    _build_heatmap_template(wb, sheet_name=SHEET_T)

    # 5. Build Mapping Sheet (from embedded data)
    _build_mapping_sheet(wb)

    # 6. Build AVL-Odriv Mapping (from embedded data)
    _build_avl_odriv_mapping(wb)

    return wb


def _copy_sheet(src_ws, dst_wb: openpyxl.Workbook, name: str):
    """Copy an entire worksheet into dst_wb with values and formatting."""
    dst_ws = dst_wb.create_sheet(name)
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row,
                                 min_col=1, max_col=src_ws.max_column):
        for cell in row:
            dst_cell = dst_ws.cell(cell.row, cell.column)
            dst_cell.value = cell.value
            if cell.has_style:
                dst_cell.font = copy.copy(cell.font)
                dst_cell.fill = copy.copy(cell.fill)
                dst_cell.alignment = copy.copy(cell.alignment)
                dst_cell.border = copy.copy(cell.border)
                dst_cell.number_format = cell.number_format
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
    for row_num, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dim.height
    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))


def _build_heatmap_template(wb: openpyxl.Workbook, sheet_name: str = TEMPLATE_SHEET):
    """Create the HeatMap Template sheet from embedded data."""
    ws = wb.create_sheet(sheet_name)

    # Row 1: header label
    ws.cell(1, 4).value = TARGET_VEHICLE_HEADER
    ws.cell(1, 4).font = Font(name="Arial", size=16)

    # Row 2: "Operation Modes" + Vehicle placeholders + Status + Comments
    ws.cell(2, 2).value = ANCHOR_TEXT
    ws.cell(2, 2).font = Font(bold=True)
    for vc in VEHICLE_COLS:
        ws.cell(2, vc).value = "Vehicle"
    ws.cell(2, 18).value = "Status"
    ws.cell(2, 19).value = "Comments"

    # Row 3: DR markers
    for vc in VEHICLE_COLS:
        ws.cell(3, vc).value = "DR"

    # Data rows
    for row_num, op_code, name, is_bold in HEATMAP_TEMPLATE_ROWS:
        ws.cell(row_num, 1).value = op_code
        ws.cell(row_num, 2).value = name
        if is_bold:
            ws.cell(row_num, 2).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for vc in VEHICLE_COLS:
        ws.column_dimensions[get_column_letter(vc)].width = 10
    ws.column_dimensions["R"].width = 12
    ws.column_dimensions["S"].width = 25


def _build_mapping_sheet(wb: openpyxl.Workbook):
    """Create the Mapping Sheet from embedded data."""
    ws = wb.create_sheet("Mapping Sheet")
    for i, (code, name) in enumerate(MAPPING_SHEET_DATA, 1):
        ws.cell(i, 1).value = code
        ws.cell(i, 2).value = name


def _build_avl_odriv_mapping(wb: openpyxl.Workbook):
    """Create the AVL-Odriv Mapping sheet from embedded data."""
    ws = wb.create_sheet("AVL-Odriv Mapping")
    for i, (code, name) in enumerate(AVL_ODRIV_MAPPING_DATA, 1):
        ws.cell(i, 1).value = code
        ws.cell(i, 2).value = name


# ═══════════════════════════════════════════════════════════════════════════════
#  COLOUR / STATUS LOGIC  (mirrors Evaluation.bas)
# ═══════════════════════════════════════════════════════════════════════════════

def get_p1_status_from_color(cell) -> str:
    """Determine P1 status from the cell's font colour (indexed or RGB).

    Mapping:
        indexed 17 / #008000  -> GREEN
        indexed 13 / #FFFF00  -> YELLOW
        indexed 10 / #FF0000  -> RED
        #FFFFFF / white       -> N/A
    """
    hex_color = _resolve_font_color_hex(cell)
    r, g, b = _hex_to_rgb(hex_color)

    fill_hex = _resolve_fill_color_hex(cell)
    fr, fg, fb = _hex_to_rgb(fill_hex)

    # Fill colour checks
    if _is_near(fr, fg, fb, 0, 176, 80) or _is_near(fr, fg, fb, 0, 158, 71):
        return "GREEN"
    if _is_near(fr, fg, fb, 255, 192, 0) or _is_near(fr, fg, fb, 255, 217, 102, 60):
        return "YELLOW"
    if _is_near(fr, fg, fb, 255, 0, 0) or _is_near(fr, fg, fb, 192, 0, 0):
        return "RED"

    # Font colour checks
    if _is_near(r, g, b, 0, 128, 0, 5):
        return "GREEN"
    if _is_near(r, g, b, 0, 176, 80) or _is_near(r, g, b, 0, 158, 71):
        return "GREEN"
    if _is_near(r, g, b, 255, 255, 0, 5):
        return "YELLOW"
    if _is_near(r, g, b, 255, 192, 0) or _is_near(r, g, b, 255, 217, 102, 60):
        return "YELLOW"
    if _is_near(r, g, b, 227, 225, 0, 30):
        return "YELLOW"
    if _is_near(r, g, b, 255, 0, 0) or _is_near(r, g, b, 192, 0, 0):
        return "RED"

    return "N/A"


def bench_diff(target_val: float, tested_val: float) -> float:
    """Return abs difference; 999 sentinel when target is zero."""
    if target_val == 0 and tested_val == 0:
        return 999.0
    if target_val == 0:
        return 999.0
    return abs(tested_val - target_val)


def evaluate_status(avl: float, p1: str, bdiff: float,
                    target_val: float, tested_val: float) -> str:
    """Mirror EvaluateStatus VBA function."""
    p1u = p1.upper().strip()

    if p1u == "N/A":
        return "N/A"
    if avl < 7 or p1u == "RED":
        return "RED"
    if avl >= 7 and p1u == "YELLOW":
        return "YELLOW"

    # AVL >= 7 and P1 == GREEN
    if bdiff == 999:
        return "GREEN"
    if tested_val >= target_val:
        return "GREEN"
    if (target_val - tested_val) <= 2:
        return "GREEN"
    return "YELLOW"


def combine_status(driv: str, resp: str) -> str:
    """Mirror CombineStatus VBA function."""
    d = driv.upper().strip() if driv else "N/A"
    r = resp.upper().strip() if resp else "N/A"
    if d == "":
        d = "N/A"
    if r == "":
        r = "N/A"

    if d == "RED" or r == "RED":
        return "RED"
    if d == "YELLOW" or r == "YELLOW":
        return "YELLOW"
    if d == "GREEN" and r == "GREEN":
        return "GREEN"
    if (d == "GREEN" and r == "N/A") or (d == "N/A" and r == "GREEN"):
        return "GREEN"
    return "N/A"


def color_cell(ws, row: int, col: int, status: str):
    """Apply fill + font colour to a results cell."""
    cell = ws.cell(row, col)
    s = status.upper().strip()
    if s == "GREEN":
        cell.fill = FILL_GREEN
        cell.font = FONT_WHITE
    elif s == "YELLOW":
        cell.fill = FILL_YELLOW
        cell.font = FONT_BLACK
    elif s == "RED":
        cell.fill = FILL_RED
        cell.font = FONT_WHITE
    else:
        cell.fill = PatternFill(fill_type=None)
        cell.font = FONT_BLACK


# ═══════════════════════════════════════════════════════════════════════════════
#  1.  HEATMAP REFRESH  (mirrors HeatMap.bas -> RefreshHeatmap)
# ═══════════════════════════════════════════════════════════════════════════════

def _collect_dest_vehicle_cols(ws, anc_row: int, anc_col: int) -> List[int]:
    """Collect destination vehicle columns by DR marker in row anc_row+1."""
    last_c = ws.max_column
    out: List[int] = []
    for c in range(anc_col + 1, last_c + 1):
        v = _cell_val(ws, anc_row + 1, c)
        if isinstance(v, str) and v.strip().upper().startswith("DR"):
            out.append(c)
    if out:
        return out
    for c in range(anc_col + 1, last_c + 1):
        v = _trim(_cell_val(ws, anc_row, c))
        if v.upper() == "COMMENTS":
            break
        if v:
            out.append(c)
        elif out:
            break
    return out


def _collect_headers(ws, anc_row: int, anc_col: int) -> List[str]:
    out: List[str] = []
    for c in range(anc_col + 1, ws.max_column + 1):
        v = _trim(_cell_val(ws, anc_row, c))
        if v:
            out.append(v)
    return out


def _collect_header_cols(ws, anc_row: int, anc_col: int) -> List[int]:
    out: List[int] = []
    for c in range(anc_col + 1, ws.max_column + 1):
        v = _trim(_cell_val(ws, anc_row, c))
        if v:
            out.append(c)
    return out


def _collect_row_labels(ws, anc_row: int, anc_col: int) -> List[str]:
    out: List[str] = []
    empty_run = 0
    last_r = ws.max_row
    for r in range(anc_row + 2, last_r + 1):
        v = _trim(_cell_val(ws, r, anc_col))
        if v:
            out.append(v)
            empty_run = 0
        else:
            empty_run += 1
            if empty_run >= 10:
                break
    return out


def _build_mode_index(ws, anc_row: int, anc_col: int) -> Dict[str, int]:
    d: Dict[str, int] = {}
    last_r = ws.max_row
    for r in range(anc_row + 2, last_r + 1):
        v = _trim(_cell_val(ws, r, anc_col))
        if v and v.lower() not in {k.lower() for k in d}:
            d[v] = r
    return d


def refresh_heatmap(wb) -> str:
    """Transfer data from Data Transfer Sheet to HeatMap Sheet."""
    ws_t = wb[SHEET_T]
    ws_s = wb[SHEET_S]

    t_anc = _find_anchor(ws_t, ANCHOR_TEXT)
    s_anc = _find_anchor(ws_s, ANCHOR_TEXT)
    if t_anc is None or s_anc is None:
        return f"Anchor text '{ANCHOR_TEXT}' not found in one or both sheets."

    t_row, t_col = t_anc
    s_row, s_col = s_anc

    t_veh_cols = _collect_dest_vehicle_cols(ws_t, t_row, t_col)
    t_modes = _collect_row_labels(ws_t, t_row, t_col)
    s_veh_hdr = _collect_headers(ws_s, s_row, s_col)
    s_veh_col = _collect_header_cols(ws_s, s_row, s_col)
    s_mode_ix = _build_mode_index(ws_s, s_row, s_col)

    if not t_veh_cols or not s_veh_col or not t_modes:
        return "No data found to process."

    n = min(len(t_veh_cols), len(s_veh_hdr))
    warnings: List[str] = []
    if len(s_veh_hdr) > len(t_veh_cols):
        warnings.append(
            f"Data Transfer Sheet has {len(s_veh_hdr)} vehicles but HeatMap "
            f"can accommodate {len(t_veh_cols)}. Only first {n} transferred."
        )

    # Vehicle header labels
    if n > 0:
        ws_t.cell(t_row - 1, t_veh_cols[0]).value = TARGET_VEHICLE_HEADER
        ws_t.cell(t_row - 1, t_veh_cols[0]).font = Font(name="Arial", size=16)
        ws_t.cell(t_row - 1, t_veh_cols[0]).alignment = Alignment(horizontal="center")

        ws_t.cell(t_row - 1, t_veh_cols[n - 1]).value = TESTED_VEHICLE_HEADER
        ws_t.cell(t_row - 1, t_veh_cols[n - 1]).font = Font(name="Arial", size=16)
        ws_t.cell(t_row - 1, t_veh_cols[n - 1]).alignment = Alignment(horizontal="center")

    # Vehicle names
    for i in range(n):
        ws_t.cell(t_row, t_veh_cols[i]).value = s_veh_hdr[i]

    # Clear old data
    last_r = t_row + 1 + len(t_modes)
    for j in range(n):
        for r in range(t_row + 2, last_r + 1):
            ws_t.cell(r, t_veh_cols[j]).value = None

    # Fill data
    filled = 0
    for i, mode in enumerate(t_modes):
        matched_key = None
        for k in s_mode_ix:
            if k.lower() == mode.lower():
                matched_key = k
                break
        if matched_key is not None:
            r_s = s_mode_ix[matched_key]
            r_t = t_row + 1 + (i + 1)
            for j in range(n):
                v = _cell_val(ws_s, r_s, s_veh_col[j])
                if _has_value(v):
                    try:
                        ws_t.cell(r_t, t_veh_cols[j]).value = float(v)
                        filled += 1
                    except (ValueError, TypeError):
                        pass

    # Hide rows missing last vehicle data
    if not DELETE_EMPTY:
        _hide_rows_missing_last_vehicle(ws_t, t_row, t_col, t_veh_cols[n - 1])
    else:
        _delete_rows_missing_last_vehicle(ws_t, t_row, t_col, t_veh_cols[n - 1])

    msg = f"HeatMap refreshed: {filled} values transferred for {n} vehicle(s)."
    if warnings:
        msg += "\n" + "\n".join(warnings)
    return msg


def _hide_rows_missing_last_vehicle(ws, anc_row: int, anc_col: int, last_veh_col: int):
    last_r = ws.max_row
    for r in range(anc_row + 2, last_r + 1):
        v = _trim(_cell_val(ws, r, anc_col))
        veh = _cell_val(ws, r, last_veh_col)
        if v == "" or not _has_value(veh):
            ws.row_dimensions[r].hidden = True
        else:
            ws.row_dimensions[r].hidden = False


def _delete_rows_missing_last_vehicle(ws, anc_row: int, anc_col: int, last_veh_col: int):
    last_r = ws.max_row
    for r in range(last_r, anc_row + 1, -1):
        v = _trim(_cell_val(ws, r, anc_col))
        veh = _cell_val(ws, r, last_veh_col)
        if v == "" or not _has_value(veh):
            ws.delete_rows(r)


# ═══════════════════════════════════════════════════════════════════════════════
#  2.  RESET  (mirrors Reset.bas)
# ═══════════════════════════════════════════════════════════════════════════════

def reset_heatmap(wb) -> str:
    """Reset the HeatMap Sheet by rebuilding from embedded template data."""
    if SHEET_T in wb.sheetnames:
        idx = wb.sheetnames.index(SHEET_T)
        del wb[SHEET_T]
    else:
        idx = len(wb.sheetnames)

    _build_heatmap_template(wb, sheet_name=SHEET_T)
    ws = wb[SHEET_T]
    wb.move_sheet(ws, offset=idx - len(wb.sheetnames) + 1)

    return "HeatMap Sheet has been reset from template."


# ═══════════════════════════════════════════════════════════════════════════════
#  3.  EVALUATION  (mirrors Evaluation.bas -> EvaluateAVLStatus)
# ═══════════════════════════════════════════════════════════════════════════════

def get_available_car_names(ws) -> List[str]:
    """Scan row 2 of Sheet1 from col H onwards for unique car names."""
    names: List[str] = []
    last_col = ws.max_column
    skip_words = {"status", "p1", "p2", "p3", "lowest events", "current status"}
    for col in range(CAR_DATA_START_COL, last_col + 1):
        name = _trim(_cell_val(ws, 2, col))
        if not name:
            continue
        if any(w in name.lower() for w in skip_words):
            continue
        if name not in names:
            names.append(name)
    return names


def _find_car_column(ws, car_name: str, start_col: int = CAR_DATA_START_COL) -> int:
    """Find column for a car name in row 2 from start_col onwards."""
    for col in range(start_col, ws.max_column + 1):
        if _trim(_cell_val(ws, 2, col)) == car_name.strip():
            return col
    return 0


def _get_tested_avl(ws_heatmap, op_code, tested_car_name: str) -> float:
    """Look up Tested AVL score from HeatMap Sheet."""
    op_key = _trim(str(op_code))

    # Find the column under "Tested Vehicle" header in row 1
    avl_col = 0
    for col in range(1, ws_heatmap.max_column + 1):
        v = _trim(_cell_val(ws_heatmap, 1, col))
        if v and "tested vehicle" in v.lower():
            avl_col = col
            break

    # Fallback: search row 2 for tested car name
    if avl_col == 0:
        for col in range(1, ws_heatmap.max_column + 1):
            if _trim(_cell_val(ws_heatmap, 2, col)) == tested_car_name.strip():
                avl_col = col
                break

    # Default to last vehicle column
    if avl_col == 0:
        avl_col = VEHICLE_COLS[-1]

    # Search column A for op code
    for r in range(1, ws_heatmap.max_row + 1):
        v = _cell_val(ws_heatmap, r, 1)
        if v is not None and _trim(str(v)) == op_key:
            return _to_float(_cell_val(ws_heatmap, r, avl_col))

    # Numeric match
    if _is_numeric(op_key):
        op_num = int(float(op_key))
        for r in range(1, ws_heatmap.max_row + 1):
            v = _cell_val(ws_heatmap, r, 1)
            if v is not None:
                try:
                    if int(float(v)) == op_num:
                        return _to_float(_cell_val(ws_heatmap, r, avl_col))
                except (ValueError, TypeError):
                    pass

    return 0.0


def evaluate_avl_status(wb, target_car: str, tested_car: str) -> str:
    """Run the full evaluation and write to Evaluation Results sheet."""
    ws1 = wb[SHEET1]
    ws_hm = wb[SHEET_T]

    target_col = _find_car_column(ws1, target_car)
    tested_col = _find_car_column(ws1, tested_car)

    if target_col == 0 or tested_col == 0:
        return (
            f"Could not find data columns for selected cars.\n"
            f"Target: {target_car} (col {target_col})\n"
            f"Tested: {tested_car} (col {tested_col})"
        )

    # Responsiveness section starts at col 12
    target_resp_col = _find_car_column(ws1, target_car, 12)
    tested_resp_col = _find_car_column(ws1, tested_car, 12)

    if target_resp_col == 0 or tested_resp_col == 0:
        return "Could not find responsiveness columns for selected cars."

    if EVAL_RESULTS in wb.sheetnames:
        del wb[EVAL_RESULTS]

    ws_r = wb.create_sheet(EVAL_RESULTS)

    headers = [
        "Op Code", "Operation", "Tested AVL",
        "Driv P1", f"Driv Target ({target_car})", f"Driv Tested ({tested_car})", "Driv Status",
        "Resp P1", f"Resp Target ({target_car})", f"Resp Tested ({tested_car})", "Resp Status",
        "Final Status",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_r.cell(1, col_idx, h)
        cell.font = FONT_BOLD_WHITE
        cell.fill = FILL_HEADER

    last_row = max(
        _last_data_row(ws1, 1),
        _last_data_row(ws1, 2),
        _last_data_row(ws1, 3),
    )

    out_row = 2
    for i in range(5, last_row + 1):
        op_code = _cell_val(ws1, i, 2)  # Column B

        if op_code is None or not _is_numeric(op_code):
            continue

        tested_avl = _get_tested_avl(ws_hm, op_code, tested_car)

        driv_p1 = get_p1_status_from_color(ws1.cell(i, 6))   # Column F
        resp_p1 = get_p1_status_from_color(ws1.cell(i, 12))  # Column L

        driv_target = _to_float(_cell_val(ws1, i, target_col))
        driv_tested = _to_float(_cell_val(ws1, i, tested_col))
        resp_target = _to_float(_cell_val(ws1, i, target_resp_col))
        resp_tested = _to_float(_cell_val(ws1, i, tested_resp_col))

        driv_bdiff = bench_diff(driv_target, driv_tested)
        resp_bdiff = bench_diff(resp_target, resp_tested)

        driv_status = evaluate_status(tested_avl, driv_p1, driv_bdiff, driv_target, driv_tested)
        resp_status = evaluate_status(tested_avl, resp_p1, resp_bdiff, resp_target, resp_tested)
        final_status = combine_status(driv_status, resp_status)

        operation_name = _cell_val(ws1, i, 3)  # Column C

        ws_r.cell(out_row, 1).value = op_code
        ws_r.cell(out_row, 2).value = operation_name
        ws_r.cell(out_row, 3).value = tested_avl
        ws_r.cell(out_row, 4).value = driv_p1
        ws_r.cell(out_row, 5).value = driv_target
        ws_r.cell(out_row, 6).value = driv_tested
        ws_r.cell(out_row, 7).value = driv_status
        ws_r.cell(out_row, 8).value = resp_p1
        ws_r.cell(out_row, 9).value = resp_target
        ws_r.cell(out_row, 10).value = resp_tested
        ws_r.cell(out_row, 11).value = resp_status
        ws_r.cell(out_row, 12).value = final_status

        color_cell(ws_r, out_row, 7, driv_status)
        color_cell(ws_r, out_row, 11, resp_status)
        color_cell(ws_r, out_row, 12, final_status)

        out_row += 1

    for col_idx in range(1, 13):
        ws_r.column_dimensions[get_column_letter(col_idx)].width = 20

    _build_overall_status(ws_r)

    return (
        f"Evaluation complete!\n"
        f"Target: {target_car}\n"
        f"Tested: {tested_car}\n"
        f"Results: {out_row - 2} rows written to '{EVAL_RESULTS}'."
    )


def _build_overall_status(ws_r):
    """Build 'Overall Status by Op Code' summary at bottom of results."""
    last_row = _last_data_row(ws_r, 1)

    codes: Dict[str, dict] = OrderedDict()
    for i in range(2, last_row + 1):
        code = _trim(_cell_val(ws_r, i, 1))
        if not code:
            continue
        status = _trim(_cell_val(ws_r, i, 12))
        name = _trim(_cell_val(ws_r, i, 2))

        if code not in codes:
            codes[code] = {"name": name, "statuses": [status]}
        else:
            codes[code]["statuses"].append(status)

    start_row = last_row + 2

    ws_r.cell(start_row, 1).value = "Overall Status by Op Code"
    ws_r.cell(start_row, 1).font = Font(bold=True)
    ws_r.cell(start_row, 1).fill = FILL_SUMMARY
    ws_r.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=4,
    )

    ws_r.cell(start_row + 1, 1).value = "Op Code"
    ws_r.cell(start_row + 1, 2).value = "Operation"
    ws_r.cell(start_row + 1, 3).value = "Overall Status"
    for c in range(1, 4):
        ws_r.cell(start_row + 1, c).font = Font(bold=True)

    r = start_row + 2
    for code, info in codes.items():
        any_red = False
        all_green = True
        has_valid = False

        for s in info["statuses"]:
            su = s.upper().strip()
            if su and su != "N/A":
                has_valid = True
                if su == "RED":
                    any_red = True
                if su != "GREEN":
                    all_green = False

        if not has_valid:
            overall = "N/A"
        elif any_red:
            overall = "RED"
        elif all_green:
            overall = "GREEN"
        else:
            overall = "YELLOW"

        ws_r.cell(r, 1).value = code
        ws_r.cell(r, 2).value = info["name"]
        ws_r.cell(r, 3).value = overall
        color_cell(ws_r, r, 3, overall)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
#  4.  SUB-OPERATION STATUS  (mirrors Updatesuboperationstatus.bas)
# ═══════════════════════════════════════════════════════════════════════════════

def update_sub_operation_heatmap(wb) -> str:
    """Read Evaluation Results, write coloured dots to HeatMap Sheet col R."""
    if EVAL_RESULTS not in wb.sheetnames:
        return "Evaluation Results sheet not found. Run Evaluation first."

    ws_eval = wb[EVAL_RESULTS]
    ws_heat = wb[SHEET_T]

    d: Dict[str, str] = {}
    last_eval = _last_data_row(ws_eval, 1)
    for i in range(2, last_eval + 1):
        code = _trim(_cell_val(ws_eval, i, 1))
        if code:
            status = _trim(_cell_val(ws_eval, i, 3)).upper()
            if not status:
                status = _trim(_cell_val(ws_eval, i, 12)).upper()
            d[code] = status

    last_heat = _last_data_row(ws_heat, 1)
    updated = 0
    for i in range(2, last_heat + 1):
        cell_b = ws_heat.cell(i, 2)
        if cell_b.font and cell_b.font.bold:
            continue

        op_code = _trim(_cell_val(ws_heat, i, 1))
        if not op_code:
            continue

        target_cell = ws_heat.cell(i, 18)  # Column R
        target_cell.value = None

        if op_code in d:
            status = d[op_code]
            if status in ("RED", "YELLOW", "GREEN"):
                target_cell.value = BULLET
                target_cell.alignment = Alignment(horizontal="center")

                if status == "RED":
                    target_cell.font = Font(size=14, color="FF0000")
                elif status == "YELLOW":
                    target_cell.font = Font(size=14, color="E3E100")
                elif status == "GREEN":
                    target_cell.font = Font(size=14, color="00B050")
                updated += 1

    return f"Sub-operation HeatMap updated: {updated} status dots written."


# ═══════════════════════════════════════════════════════════════════════════════
#  5.  OPERATION MODE STATUS  (mirrors OperationModeStatus.bas)
# ═══════════════════════════════════════════════════════════════════════════════

def update_operation_mode_status(wb, status_col: int = 18) -> str:
    """Aggregate sub-operation statuses into group header rows."""
    ws = wb[SHEET_T]
    last_row = _last_data_row(ws, 2)

    grp_start = 0
    groups_updated = 0

    for i in range(2, last_row + 2):
        cell_b = ws.cell(i, 2) if i <= last_row else None
        is_bold = cell_b is not None and cell_b.font and cell_b.font.bold

        if i > last_row or is_bold:
            if grp_start != 0 and grp_start - 1 > 2:
                _evaluate_group_status(ws, grp_start, i - 1, status_col)
                groups_updated += 1
            if i <= last_row:
                grp_start = i + 1

    return f"Operation mode statuses updated: {groups_updated} groups evaluated."


def _evaluate_group_status(ws, start_row: int, end_row: int, status_col: int):
    """Evaluate a single group's status and write to header cell."""
    if start_row - 1 <= 2:
        return

    red_cnt = 0
    yellow_cnt = 0
    total_cnt = 0

    for r in range(start_row, end_row + 1):
        cell = ws.cell(r, status_col)
        if cell.value and str(cell.value).strip():
            total_cnt += 1
            fc = _resolve_font_color_hex(cell)
            red, green, blue = _hex_to_rgb(fc)
            if _is_near(red, green, blue, 255, 0, 0, 45):
                red_cnt += 1
            elif (_is_near(red, green, blue, 227, 225, 0, 45) or
                  _is_near(red, green, blue, 255, 255, 0, 45)):
                yellow_cnt += 1

    if total_cnt == 0:
        return

    header_cell = ws.cell(start_row - 1, status_col)
    header_cell.font = Font(bold=True, color="000000")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    if red_cnt > 0:
        header_cell.value = "NOK"
        header_cell.fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid")
    elif total_cnt > 0 and yellow_cnt / total_cnt > 0.35:
        header_cell.value = "Acceptable"
        header_cell.fill = PatternFill(
            start_color="E3E100", end_color="E3E100", fill_type="solid")
    else:
        header_cell.value = "OK"
        header_cell.fill = PatternFill(
            start_color="00B050", end_color="00B050", fill_type="solid")


# ═══════════════════════════════════════════════════════════════════════════════
#  6.  CLEAR ALL  (mirrors Clearall.bas)
# ═══════════════════════════════════════════════════════════════════════════════

def clear_sheet1(wb) -> str:
    """Clear all cells in Sheet1."""
    if SHEET1 not in wb.sheetnames:
        return f"Sheet '{SHEET1}' not found."
    ws = wb[SHEET1]
    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    return "Sheet1 cleared."


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_sheet_data(wb, sheet_name: str) -> bytes:
    """Export a sheet's visible data to a new XLSX file (in memory)."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    ws = wb[sheet_name]
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name

    out_r = 1
    for r in range(1, ws.max_row + 1):
        if ws.row_dimensions[r].hidden:
            continue
        out_c = 1
        for c in range(1, ws.max_column + 1):
            col_letter = get_column_letter(c)
            if ws.column_dimensions[col_letter].hidden:
                continue
            src = ws.cell(r, c)
            dst = new_ws.cell(out_r, out_c)
            dst.value = src.value
            if src.has_style:
                dst.font = copy.copy(src.font)
                dst.fill = copy.copy(src.fill)
                dst.alignment = copy.copy(src.alignment)
                dst.border = copy.copy(src.border)
                dst.number_format = src.number_format
            out_c += 1
        out_r += 1

    buf = io.BytesIO()
    new_wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA PREVIEW
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_to_dataframe(ws, max_rows: int = 500):
    """Convert a worksheet to a pandas DataFrame for display."""
    import pandas as pd

    data: List[List[Any]] = []
    for r in range(1, min(ws.max_row + 1, max_rows + 1)):
        row_data: List[Any] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            row_data.append(v)
        data.append(row_data)

    if not data:
        return pd.DataFrame()

    headers = [str(v) if v else f"Col{i+1}" for i, v in enumerate(data[0])]
    df = pd.DataFrame(data[1:], columns=headers)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT GUI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="AVL-DRIVE Heatmap Tool",
        page_icon="\U0001f527",
        layout="wide",
    )

    st.title("\U0001f527 AVL-DRIVE Heatmap Tool \u2014 Python Edition")
    st.caption("Fully independent Python replica of the Excel VBA tool (v5.1)")

    # ── Sidebar: Mode Selection & File Upload ────────────────────────────
    st.sidebar.header("\U0001f4c2 Data Input")

    mode = st.sidebar.radio(
        "Input Mode",
        [
            "\U0001f4c1 Upload individual data files (standalone)",
            "\U0001f4e6 Upload .xlsm workbook (legacy)",
        ],
        index=0,
        help=(
            "Standalone mode: upload your Data Transfer Sheet and Sheet1 files.\n"
            "Legacy mode: upload the original .xlsm file."
        ),
    )

    if "standalone" in mode.lower():
        _render_standalone_mode()
    else:
        _render_legacy_mode()


def _render_standalone_mode():
    """Standalone mode: upload Data Transfer Sheet + Sheet1 separately."""
    st.sidebar.markdown("---")
    st.sidebar.subheader("\u2460 Data Transfer Sheet")
    st.sidebar.caption(
        "Upload the AVL-DRIVE scores file.\n"
        "Format: Col A = Op Code, Col B = Operation Name, "
        "then vehicle scores in alternating columns (D, F, H, J\u2026).\n"
        "Row 1 = car names, Row 2 = 'DR' markers."
    )
    dt_file = st.sidebar.file_uploader(
        "Data Transfer Sheet",
        type=["xlsx", "xls", "csv"],
        key="dt_upload",
    )

    st.sidebar.markdown("---")
    st.sidebar.subheader("\u2461 Sheet1 (Benchmark Data)")
    st.sidebar.caption(
        "Upload the benchmark + P1 status file (.xlsx to preserve colours).\n"
        "Format: Op codes in col B, operation names in col C, "
        "coloured P1 dots in cols F & L, benchmark values in car columns."
    )
    s1_file = st.sidebar.file_uploader(
        "Sheet1 (benchmark data)",
        type=["xlsx", "xls"],
        key="s1_upload",
    )

    if dt_file is None or s1_file is None:
        st.info(
            "\U0001f446 Upload both data files in the sidebar to get started."
        )
        _render_format_guide()
        return

    # Build workbook when new files are uploaded
    dt_key = f"dt_{dt_file.name}_{dt_file.size}"
    s1_key = f"s1_{s1_file.name}_{s1_file.size}"

    need_rebuild = (
        st.session_state.get("dt_key") != dt_key
        or st.session_state.get("s1_key") != s1_key
        or "wb_bytes" not in st.session_state
    )

    if need_rebuild:
        st.session_state["dt_key"] = dt_key
        st.session_state["s1_key"] = s1_key
        st.session_state["messages"] = []

        dt_bytes = dt_file.read()
        s1_bytes = s1_file.read()

        if dt_file.name.endswith(".csv"):
            import pandas as pd
            df = pd.read_csv(io.BytesIO(dt_bytes))
            dt_wb = openpyxl.Workbook()
            ws = dt_wb.active
            for c_idx, col_name in enumerate(df.columns, 1):
                ws.cell(1, c_idx).value = col_name
            for r_idx, row_vals in df.iterrows():
                for c_idx, val in enumerate(row_vals, 1):
                    ws.cell(int(r_idx) + 2, c_idx).value = val
        else:
            dt_wb = openpyxl.load_workbook(
                io.BytesIO(dt_bytes), data_only=False)

        s1_wb = openpyxl.load_workbook(
            io.BytesIO(s1_bytes), data_only=False)

        wb = build_workbook_from_uploads(dt_wb, s1_wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.session_state["wb_bytes"] = buf.read()
        st.session_state["wb_name"] = "avl_heatmap_output.xlsx"

        dt_wb.close()
        s1_wb.close()

    st.sidebar.success(
        f"\u2705 Data loaded:\n"
        f"\u2022 DT: `{dt_file.name}`\n"
        f"\u2022 S1: `{s1_file.name}`"
    )

    _render_tool_ui()


def _render_legacy_mode():
    """Legacy mode: upload the original .xlsm file."""
    st.sidebar.markdown("---")
    uploaded = st.sidebar.file_uploader(
        "Upload .xlsm workbook",
        type=["xlsm", "xlsx"],
        key="xlsm_upload",
        help="Upload the AVLDrive_Heatmap_Tool .xlsm file",
    )

    if uploaded is not None:
        if ("wb_bytes" not in st.session_state
                or st.session_state.get("wb_name") != uploaded.name):
            st.session_state["wb_bytes"] = uploaded.read()
            st.session_state["wb_name"] = uploaded.name
            st.session_state["messages"] = []

    if "wb_bytes" not in st.session_state:
        st.info("\U0001f446 Upload an `.xlsm` workbook to get started.")
        _render_format_guide()
        return

    st.sidebar.success(f"\u2705 Loaded: `{st.session_state['wb_name']}`")
    _render_tool_ui()


def _render_format_guide():
    """Show format guides for both input modes."""
    st.markdown("""
    ### How It Works

    This tool is a **complete standalone Python replica** of the Excel VBA-based
    AVL-DRIVE Heatmap Tool. It does **not** require the original `.xlsm` file.

    ---

    ### \U0001f4c1 Standalone Mode \u2014 Upload Your Data Files

    Upload two files that you normally paste into the Excel tool:

    #### \u2460 Data Transfer Sheet (CSV or Excel)
    The AVL-DRIVE scores per vehicle. Format:

    | Col A | Col B | Col C | Col D | Col E | Col F | ... |
    |-------|-------|-------|-------|-------|-------|-----|
    | | Operation Modes | | *Car Name 1* | | *Car Name 2* | ... |
    | | Operation Modes | | DR | | DR | ... |
    | 10000000 | AVL-DRIVE Rating | | 7.6 | | 7.4 | ... |
    | 10100000 | Drive away | | 7.7 | | 7.8 | ... |
    | ... | ... | | ... | | ... | ... |

    - **Col A**: Op code (numeric)
    - **Col B**: Operation name
    - **Cols D, F, H, J**: Vehicle AVL scores (even columns, separated by empty cols)
    - **Row 1**: Car names in data columns
    - **Row 2**: "DR" markers in data columns

    #### \u2461 Sheet1 \u2014 Benchmark Data (Excel only, .xlsx)
    Must be Excel format to **preserve P1 status colours** (green/yellow/red dots).

    | Col A | Col B | Col C | Col F | Col I | Col J | Col L | Col O | Col P |
    |-------|-------|-------|-------|-------|-------|-------|-------|-------|
    | | | | Drivability | | | Responsiveness | | |
    | | | | Current Status | *Car 1* | *Car 2* | Current Status | *Car 1* | *Car 2* |
    | USE CASE | | | P1 | | | P1 | | |
    | Drive away | | | | 98.8 | | | 70.9 | |
    | | 10101300 | Creep Eng On | \u25cf | 100 | 73.7 | \u25cf | 100 | 100 |

    - **Col B**: Op code (numeric)
    - **Col C**: Sub-operation name
    - **Col F**: Drivability P1 status (coloured \u25cf dot)
    - **Col L**: Responsiveness P1 status (coloured \u25cf dot)
    - **Car name columns**: Benchmark percentage values
    - Car names appear in **Row 2**

    ---

    ### \U0001f4e6 Legacy Mode \u2014 Upload .xlsm

    Upload the original `AVLDrive_Heatmap_Tool version_5.1.xlsm` file directly.

    ---

    ### \u26a1 Available Operations

    | Button | Function |
    |--------|----------|
    | **HeatMap** | Transfer data from Data Transfer Sheet \u2192 HeatMap Sheet |
    | **Reset** | Restore HeatMap Sheet from built-in template |
    | **Evaluation** | Evaluate AVL statuses with car selection |
    | **Suboperation Status** | Write coloured status dots to HeatMap |
    | **Operation Mode Status** | Aggregate group statuses |
    | **Export** | Download a sheet's visible data as XLSX |
    """)


def _render_tool_ui():
    """Render the main tool UI (shared between standalone and legacy modes)."""
    wb = openpyxl.load_workbook(
        io.BytesIO(st.session_state["wb_bytes"]),
        data_only=False,
    )
    st.sidebar.write(f"Sheets: {', '.join(wb.sheetnames)}")

    if "messages" not in st.session_state:
        st.session_state["messages"] = []

    # ── Action Buttons ───────────────────────────────────────────────────
    st.subheader("\u26a1 Actions")

    col1, col2, col3, col4, col5, col6 = st.columns(6)

    with col1:
        if st.button("\U0001f5fa\ufe0f HeatMap", use_container_width=True,
                     help="Refresh heatmap from Data Transfer Sheet"):
            msg = refresh_heatmap(wb)
            _save_wb(wb)
            st.session_state["messages"].append(("info", msg))
            st.rerun()

    with col2:
        if st.button("\U0001f504 Reset", use_container_width=True,
                     help="Reset HeatMap from template"):
            msg = reset_heatmap(wb)
            _save_wb(wb)
            st.session_state["messages"].append(("info", msg))
            st.rerun()

    with col3:
        if st.button("\U0001f4ca Evaluation", use_container_width=True,
                     help="Evaluate AVL statuses"):
            st.session_state["show_eval_dialog"] = True

    with col4:
        if st.button("\U0001f535 Subop Status", use_container_width=True,
                     help="Update sub-operation status dots"):
            msg = update_sub_operation_heatmap(wb)
            _save_wb(wb)
            st.session_state["messages"].append(("info", msg))
            st.rerun()

    with col5:
        if st.button("\U0001f4c8 Op Mode Status", use_container_width=True,
                     help="Update operation mode statuses"):
            msg = update_operation_mode_status(wb)
            _save_wb(wb)
            st.session_state["messages"].append(("info", msg))
            st.rerun()

    with col6:
        if st.button("\U0001f5d1\ufe0f Clear Sheet1", use_container_width=True,
                     help="Clear all data in Sheet1"):
            msg = clear_sheet1(wb)
            _save_wb(wb)
            st.session_state["messages"].append(("info", msg))
            st.rerun()

    # ── Evaluation Dialog ────────────────────────────────────────────────
    if st.session_state.get("show_eval_dialog"):
        st.divider()
        st.subheader("\U0001f697 Car Selection for Evaluation")

        if SHEET1 not in wb.sheetnames:
            st.error(f"Sheet '{SHEET1}' not found in the workbook.")
        else:
            ws1 = wb[SHEET1]
            car_names = get_available_car_names(ws1)

            if not car_names:
                st.error(
                    "No car names found in Sheet1 row 2 "
                    "(starting from column H)."
                )
            else:
                ecol1, ecol2 = st.columns(2)
                with ecol1:
                    target_car = st.selectbox(
                        "\U0001f3af Target Car", car_names, key="eval_target")
                with ecol2:
                    tested_idx = (
                        min(1, len(car_names) - 1) if len(car_names) > 1
                        else 0
                    )
                    tested_car = st.selectbox(
                        "\U0001f52c Tested Car", car_names,
                        index=tested_idx, key="eval_tested")

                if target_car == tested_car:
                    st.warning(
                        "\u26a0\ufe0f Same car selected for both "
                        "Target and Tested."
                    )

                bcol1, bcol2 = st.columns(2)
                with bcol1:
                    if st.button("\u2705 Run Evaluation", type="primary"):
                        msg = evaluate_avl_status(wb, target_car, tested_car)
                        _save_wb(wb)
                        st.session_state["messages"].append(("success", msg))
                        st.session_state["show_eval_dialog"] = False
                        st.rerun()
                with bcol2:
                    if st.button("\u274c Cancel"):
                        st.session_state["show_eval_dialog"] = False
                        st.rerun()

    # ── Show messages ────────────────────────────────────────────────────
    for msg_type, msg_text in st.session_state.get("messages", []):
        if msg_type == "success":
            st.success(msg_text)
        elif msg_type == "error":
            st.error(msg_text)
        else:
            st.info(msg_text)

    if st.session_state.get("messages"):
        st.session_state["messages"] = []

    # ── Download ─────────────────────────────────────────────────────────
    st.divider()
    st.subheader("\U0001f4be Download")
    dcol1, dcol2 = st.columns(2)

    with dcol1:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            "\U0001f4e5 Download Modified Workbook",
            data=buf.getvalue(),
            file_name=st.session_state.get("wb_name", "output.xlsx"),
            mime=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            use_container_width=True,
        )

    with dcol2:
        export_sheet = st.selectbox(
            "Export sheet", wb.sheetnames, key="export_sheet")
        if st.button("\U0001f4e4 Export Sheet as XLSX",
                     use_container_width=True):
            try:
                data = export_sheet_data(wb, export_sheet)
                st.download_button(
                    f"\u2b07\ufe0f Download {export_sheet}.xlsx",
                    data=data,
                    file_name=f"{export_sheet}.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    ),
                )
            except Exception as e:
                st.error(f"Export error: {e}")

    # ── Sheet Preview ────────────────────────────────────────────────────
    st.divider()
    st.subheader("\U0001f4cb Sheet Preview")

    import pandas as pd
    preview_sheet = st.selectbox(
        "Select sheet to preview", wb.sheetnames, key="preview_sheet")
    ws_preview = wb[preview_sheet]

    df = sheet_to_dataframe(ws_preview, max_rows=200)
    if not df.empty:
        st.dataframe(df, use_container_width=True, height=400)
    else:
        st.write("Sheet is empty.")

    wb.close()


def _save_wb(wb):
    """Persist modified workbook back to session state."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.session_state["wb_bytes"] = buf.read()


if __name__ == "__main__":
    main()
